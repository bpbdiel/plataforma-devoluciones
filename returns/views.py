from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from .models import Appeal, Product, Return, ReturnPhoto, UserProfile
from .forms import AppealForm, PlatformPasswordChangeForm, PlatformUserCreationForm, PlatformUserUpdateForm, ReturnForm, ReturnPhotoUploadForm


def paginate_queryset(request, queryset, page_param='page', per_page=12):
    paginator = Paginator(queryset, per_page)
    return paginator.get_page(request.GET.get(page_param))


def querystring_without(request, *keys):
    query = request.GET.copy()
    for key in keys:
        query.pop(key, None)
    encoded = query.urlencode()
    return f'{encoded}&' if encoded else ''


def period_dates(request):
    today = timezone.localdate()
    period = request.GET.get('period', 'month')

    if period == 'all':
        return '', None, None
    if period == 'today':
        return period, today, today
    if period == 'week':
        return period, today - timedelta(days=today.weekday()), today
    if period == 'month':
        return period, today.replace(day=1), today
    if period == 'year':
        return period, today.replace(month=1, day=1), today
    return period, None, None


def apply_period_filter(queryset, request, field_name):
    period, start_date, end_date = period_dates(request)
    if start_date:
        queryset = queryset.filter(**{f'{field_name}__gte': start_date})
    if end_date:
        queryset = queryset.filter(**{f'{field_name}__lte': end_date})
    return queryset, period


def must_change_password(user):
    if not user.is_authenticated:
        return False
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile.force_password_change


def filtered_report_returns(request):
    devoluciones = Return.objects.select_related('creado_por').prefetch_related(
        'apelaciones__creado_por',
        'apelaciones__actualizado_por',
        'fotos',
    )
    q = request.GET.get('q', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    seller = request.GET.get('seller', '').strip()
    estado = request.GET.get('estado', '').strip()
    condicion = request.GET.get('condicion', '').strip()
    grado = request.GET.get('grado', '').strip()
    apelacion = request.GET.get('apelacion', '').strip()
    appeal_status = request.GET.get('appeal_status', '').strip()
    period, start_date, end_date = period_dates(request)

    if q:
        devoluciones = devoluciones.filter(
            Q(numero_orden__icontains=q) |
            Q(seller__icontains=q) |
            Q(sku__icontains=q) |
            Q(ean__icontains=q) |
            Q(producto_nombre__icontains=q) |
            Q(marca__icontains=q) |
            Q(categoria__icontains=q) |
            Q(apelaciones__numero_apelacion__icontains=q) |
            Q(apelaciones__detalle__icontains=q) |
            Q(apelaciones__estado_cuenta__icontains=q) |
            Q(apelaciones__monto_devuelto__icontains=q)
        )
    if start_date:
        devoluciones = devoluciones.filter(fecha_ingreso__gte=start_date)
    elif fecha_desde:
        devoluciones = devoluciones.filter(fecha_ingreso__gte=fecha_desde)
    if end_date:
        devoluciones = devoluciones.filter(fecha_ingreso__lte=end_date)
    elif fecha_hasta:
        devoluciones = devoluciones.filter(fecha_ingreso__lte=fecha_hasta)
    if seller:
        devoluciones = devoluciones.filter(seller=seller)
    if estado:
        devoluciones = devoluciones.filter(estado=estado)
    if condicion:
        devoluciones = devoluciones.filter(condicion_producto=condicion)
    if grado:
        devoluciones = devoluciones.filter(grado=grado)
    if apelacion == 'con':
        devoluciones = devoluciones.filter(apelaciones__isnull=False)
    elif apelacion == 'sin':
        devoluciones = devoluciones.filter(apelaciones__isnull=True)
    if appeal_status:
        devoluciones = devoluciones.filter(apelaciones__status=appeal_status)

    return devoluciones.distinct()


def report_export_response(devoluciones):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    headers = [
        'ID devolución',
        'Fecha ingreso',
        'Fecha actualización',
        'Seller',
        'Número de orden',
        'SKU',
        'EAN',
        'Producto',
        'Marca',
        'Categoría',
        'Precio venta CLP',
        'Cantidad',
        'Ingresado a bodega',
        'Estado devolución',
        'Estado del producto',
        'Grado',
        'Sub daños',
        'Detalles del daño',
        'Notas internas',
        'Fotos',
        'Creado por',
        'Tiene apelación',
        'ID apelación',
        'N° apelación o ticket',
        'Detalle apelación',
        'Status apelación',
        'Monto CLP',
        'Diferencia CLP',
        'Creada por',
        'Actualizada por',
        'Fecha creación apelación',
        'Fecha actualización apelación',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for devolucion in devoluciones:
        apelaciones = list(devolucion.apelaciones.all())
        if not apelaciones:
            apelaciones = [None]

        for apelacion in apelaciones:
            diferencia = apelacion.diferencia_pagada if apelacion else None
            ws.append([
                devolucion.id,
                devolucion.fecha_ingreso,
                timezone.localtime(devolucion.fecha_actualizacion).strftime('%d/%m/%Y %H:%M') if devolucion.fecha_actualizacion else '',
                devolucion.get_seller_display(),
                devolucion.numero_orden,
                devolucion.sku,
                devolucion.ean,
                devolucion.producto_nombre,
                devolucion.marca,
                devolucion.categoria,
                int(devolucion.precio_venta) if devolucion.precio_venta is not None else '',
                devolucion.cantidad,
                'Sí' if devolucion.ingresado_bodega else 'No',
                devolucion.get_estado_display(),
                devolucion.get_condicion_producto_display(),
                devolucion.grado,
                devolucion.get_sub_danos_display_text(),
                devolucion.detalles_dano,
                devolucion.notas_internas,
                devolucion.fotos.count(),
                devolucion.creado_por.username if devolucion.creado_por else '',
                'Sí' if apelacion else 'No',
                apelacion.id if apelacion else '',
                apelacion.numero_apelacion if apelacion else '',
                apelacion.detalle if apelacion else '',
                apelacion.get_status_display() if apelacion else '',
                apelacion.monto_devuelto if apelacion and apelacion.status == 'pagado' else '',
                int(diferencia) if diferencia is not None else '',
                apelacion.creado_por.username if apelacion and apelacion.creado_por else '',
                apelacion.actualizado_por.username if apelacion and apelacion.actualizado_por else '',
                timezone.localtime(apelacion.creado_en).strftime('%d/%m/%Y %H:%M') if apelacion else '',
                timezone.localtime(apelacion.actualizado_en).strftime('%d/%m/%Y %H:%M') if apelacion else '',
            ])

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)
        for cell in column_cells:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = timezone.localdate().strftime('reporte_devoluciones_%Y%m%d.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if must_change_password(user):
                return redirect('password_change_required')
            return redirect('dashboard')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    else:
        form = AuthenticationForm()
    return render(request, 'returns/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def password_change_required(request):
    if request.method == 'POST':
        form = PlatformPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            UserProfile.objects.update_or_create(
                user=user,
                defaults={'force_password_change': False},
            )
            update_session_auth_hash(request, user)
            messages.success(request, 'Contraseña actualizada correctamente.')
            return redirect('dashboard')
    else:
        form = PlatformPasswordChangeForm(request.user)
    return render(request, 'returns/password_change_required.html', {'form': form})


@login_required
def product_lookup_by_ean(request):
    query = (
        request.GET.get('q', '').strip()
        or request.GET.get('sku', '').strip()
        or request.GET.get('ean', '').strip()
    )
    if not query:
        return JsonResponse({'found': False, 'error': 'Ingresa o escanea un SKU o EAN.'}, status=400)

    product = Product.objects.filter(sku__iexact=query).first()
    if not product:
        product = Product.objects.filter(ean__iexact=query).first()
    if not product:
        return JsonResponse({'found': False, 'error': 'No se encontró un producto con ese SKU o EAN.'}, status=404)

    return JsonResponse({
        'found': True,
        'product': {
            'ean': product.ean,
            'sku': product.sku,
            'nombre': product.nombre,
            'marca': product.marca,
            'categoria': product.categoria,
        },
    })


@login_required
def dashboard(request):
    today = timezone.localdate()
    period = request.GET.get('period', 'month')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if period == 'today':
        start_date = today
        end_date = today
        period_label = 'Hoy'
    elif period == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
        period_label = 'Esta semana'
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
        period_label = 'Año actual'
    elif period == 'custom':
        start_date = fecha_desde or None
        end_date = fecha_hasta or None
        period_label = 'Personalizado'
    else:
        period = 'month'
        start_date = today.replace(day=1)
        end_date = today
        period_label = 'Mes actual'

    devoluciones_qs = Return.objects.all()
    apelaciones_qs = Appeal.objects.all()
    if start_date:
        devoluciones_qs = devoluciones_qs.filter(fecha_ingreso__gte=start_date)
        apelaciones_qs = apelaciones_qs.filter(creado_en__date__gte=start_date)
    if end_date:
        devoluciones_qs = devoluciones_qs.filter(fecha_ingreso__lte=end_date)
        apelaciones_qs = apelaciones_qs.filter(creado_en__date__lte=end_date)

    def build_donut(items, total, color_map):
        if not total:
            return 'var(--gray-soft) 0deg 360deg'

        start = 0
        segments = []
        for value, count in items:
            degrees = round((count / total) * 360, 2)
            end = start + degrees
            if count:
                segments.append(f'{color_map[value]} {start}deg {end}deg')
            start = end
        return ', '.join(segments) or 'var(--gray-soft) 0deg 360deg'

    def clp_format(value):
        return f'${value:,.0f}'.replace(',', '.') + ' CLP'

    def clp_compact(value):
        value = int(value or 0)
        if abs(value) >= 1_000_000:
            compact = value / 1_000_000
            return f'${compact:.1f}M'.replace('.0M', 'M')
        if abs(value) >= 1_000:
            return f'${round(value / 1_000):.0f}K'
        return f'${value}'

    grado_counts = {
        grado: devoluciones_qs.filter(grado=grado).count()
        for grado, _ in Return.GRADO_CHOICES
    }
    status_counts = {
        status: apelaciones_qs.filter(status=status).count()
        for status, _ in Appeal.STATUS_CHOICES
    }
    total_devoluciones = devoluciones_qs.count()
    total_apelaciones = apelaciones_qs.count()
    total_valor_devoluciones = sum(
        devolucion.precio_venta * devolucion.cantidad
        for devolucion in devoluciones_qs.exclude(precio_venta__isnull=True)
    )
    apelaciones_pagadas = status_counts.get('pagado', 0)
    tasa_pago_apelaciones = round((apelaciones_pagadas / total_apelaciones) * 100) if total_apelaciones else 0
    total_pagado_apelaciones = sum(
        int(apelacion.monto_devuelto)
        for apelacion in apelaciones_qs.filter(status='pagado').only('monto_devuelto')
        if apelacion.monto_devuelto and apelacion.monto_devuelto.isdigit()
    )
    costo_devoluciones_pagadas = 0
    for apelacion in apelaciones_qs.filter(status='pagado').select_related('devolucion').only('monto_devuelto', 'devolucion__precio_venta', 'devolucion__cantidad'):
        if apelacion.monto_pagado is not None and apelacion.devolucion.valor_total is not None:
            costo_devoluciones_pagadas += int(apelacion.devolucion.valor_total)
    perdida_estimada_apelaciones = max(costo_devoluciones_pagadas - total_pagado_apelaciones, 0)
    grado_colors = {
        'A': 'var(--green)',
        'C': 'var(--yellow)',
        'D': 'var(--red)',
        'E': 'var(--purple)',
    }
    status_colors = {
        'en_proceso': 'var(--yellow)',
        'rechazado': 'var(--red)',
        'proceso_pago': 'var(--purple)',
        'pagado': 'var(--green)',
    }

    grado_chart = [
        {
            'label': label,
            'value': grado,
            'count': count,
            'percent': round((count / total_devoluciones) * 100) if total_devoluciones else 0,
        }
        for grado, label in Return.GRADO_CHOICES
        for count in [grado_counts[grado]]
    ]
    appeal_chart = [
        {
            'label': label,
            'value': status,
            'count': count,
            'percent': round((count / total_apelaciones) * 100) if total_apelaciones else 0,
        }
        for status, label in Appeal.STATUS_CHOICES
        for count in [status_counts[status]]
    ]
    raw_category_counts = (
        devoluciones_qs.exclude(categoria='')
        .values('categoria')
        .annotate(count=Count('id'))
        .order_by('-count', 'categoria')[:5]
    )
    top_category_count = raw_category_counts[0]['count'] if raw_category_counts else 0
    category_chart = [
        {
            'label': item['categoria'],
            'count': item['count'],
            'percent': round((item['count'] / total_devoluciones) * 100) if total_devoluciones else 0,
            'relative_percent': round((item['count'] / top_category_count) * 100) if top_category_count else 0,
        }
        for item in raw_category_counts
    ]
    seller_labels = dict(Return.SELLER_CHOICES)
    seller_bodega_counts = (
        devoluciones_qs.values('seller')
        .annotate(
            total=Count('id'),
            ingresados=Count('id', filter=Q(ingresado_bodega=True)),
            pendientes=Count('id', filter=Q(ingresado_bodega=False)),
        )
        .order_by('-total', 'seller')
    )
    total_ingresado_bodega = devoluciones_qs.filter(ingresado_bodega=True).count()
    total_pendiente_bodega = total_devoluciones - total_ingresado_bodega
    seller_bodega_chart = [
        {
            'label': seller_labels.get(item['seller'], item['seller']),
            'total': item['total'],
            'ingresados': item['ingresados'],
            'pendientes': item['pendientes'],
            'ingresados_percent': round((item['ingresados'] / item['total']) * 100) if item['total'] else 0,
            'pendientes_percent': round((item['pendientes'] / item['total']) * 100) if item['total'] else 0,
        }
        for item in seller_bodega_counts
    ]
    channel_totals = {}
    for devolucion in devoluciones_qs:
        channel = channel_totals.setdefault(
            devolucion.seller,
            {
                'label': seller_labels.get(devolucion.seller, devolucion.seller),
                'cost': 0,
                'returned': 0,
            },
        )
        if devolucion.precio_venta is not None:
            channel['cost'] += devolucion.precio_venta * devolucion.cantidad

    for apelacion in apelaciones_qs.filter(status='pagado').select_related('devolucion').only('monto_devuelto', 'devolucion__seller'):
        if not (apelacion.monto_devuelto and apelacion.monto_devuelto.isdigit()):
            continue
        channel = channel_totals.setdefault(
            apelacion.devolucion.seller,
            {
                'label': seller_labels.get(apelacion.devolucion.seller, apelacion.devolucion.seller),
                'cost': 0,
                'returned': 0,
            },
        )
        channel['returned'] += int(apelacion.monto_devuelto)

    max_channel_amount = max(
        (max(item['cost'], item['returned']) for item in channel_totals.values()),
        default=0,
    )
    channel_chart = [
        {
            'label': item['label'],
            'cost_display': clp_format(item['cost']),
            'cost_compact': clp_compact(item['cost']),
            'returned_display': clp_format(item['returned']),
            'returned_compact': clp_compact(item['returned']),
            'cost_percent': round((item['cost'] / max_channel_amount) * 100) if max_channel_amount else 0,
            'returned_percent': round((item['returned'] / max_channel_amount) * 100) if max_channel_amount else 0,
            'recovery_percent': round((item['returned'] / item['cost']) * 100) if item['cost'] else 0,
            'recovery_label': f"{round((item['returned'] / item['cost']) * 100)}% recuperado" if item['cost'] else 'Sin costo en periodo',
        }
        for item in sorted(channel_totals.values(), key=lambda value: (value['cost'], value['returned']), reverse=True)
    ]

    context = {
        'grado_chart': grado_chart,
        'appeal_chart': appeal_chart,
        'category_chart': category_chart,
        'channel_chart': channel_chart,
        'grade_donut_gradient': build_donut(
            [(grado, grado_counts[grado]) for grado, _ in Return.GRADO_CHOICES],
            total_devoluciones,
            grado_colors,
        ),
        'appeal_donut_gradient': build_donut(
            [(status, status_counts[status]) for status, _ in Appeal.STATUS_CHOICES],
            total_apelaciones,
            status_colors,
        ),
        'total_devoluciones': total_devoluciones,
        'total_apelaciones': total_apelaciones,
        'total_valor_devoluciones_display': clp_format(total_valor_devoluciones),
        'seller_count': len(seller_bodega_chart),
        'seller_bodega_chart': seller_bodega_chart,
        'total_ingresado_bodega': total_ingresado_bodega,
        'total_pendiente_bodega': total_pendiente_bodega,
        'apelaciones_pagadas': apelaciones_pagadas,
        'tasa_pago_apelaciones': tasa_pago_apelaciones,
        'total_pagado_apelaciones': total_pagado_apelaciones,
        'total_pagado_apelaciones_display': clp_format(total_pagado_apelaciones),
        'perdida_estimada_apelaciones': perdida_estimada_apelaciones,
        'perdida_estimada_apelaciones_display': clp_format(perdida_estimada_apelaciones),
        'period': period,
        'period_label': period_label,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'returns/home_dashboard.html', context)


@login_required
def returns_dashboard(request):
    devoluciones = Return.objects.prefetch_related('apelaciones')

    # Filtros
    q = request.GET.get('q', '')
    condicion = request.GET.get('condicion', '')
    devoluciones, period = apply_period_filter(devoluciones, request, 'fecha_ingreso')
    if q:
        devoluciones = devoluciones.filter(
            Q(numero_orden__icontains=q) |
            Q(seller__icontains=q) |
            Q(sku__icontains=q) |
            Q(ean__icontains=q) |
            Q(producto_nombre__icontains=q) |
            Q(marca__icontains=q) |
            Q(categoria__icontains=q)
        )
    if condicion:
        devoluciones = devoluciones.filter(condicion_producto=condicion)

    devoluciones_page = paginate_queryset(request, devoluciones, 'page', 12)

    def clp_format(value):
        return f'${value:,.0f}'.replace(',', '.') + ' CLP'

    devoluciones_con_precio = devoluciones.exclude(precio_venta__isnull=True)
    total_valorizado = sum(
        devolucion.precio_venta * devolucion.cantidad
        for devolucion in devoluciones_con_precio
    )

    stats = {
        'total': devoluciones.count(),
        'nuevo': devoluciones.filter(condicion_producto='nuevo').count(),
        'caja_danada': devoluciones.filter(condicion_producto='caja_danada').count(),
        'danado': devoluciones.filter(condicion_producto='danado').count(),
        'dano_estetico': devoluciones.filter(condicion_producto='dano_estetico').count(),
        'extraviado': devoluciones.filter(condicion_producto='extraviado').count(),
        'total_valorizado_display': clp_format(total_valorizado),
    }

    context = {
        'devoluciones': devoluciones_page,
        'devoluciones_page': devoluciones_page,
        'devoluciones_page_query': querystring_without(request, 'page'),
        'period': period,
        'period_query': querystring_without(request, 'period', 'page'),
        'stats': stats,
        'q': q,
        'condicion_filter': condicion,
        'condicion_choices': Return.CONDICION_CHOICES,
        'create_form': ReturnForm(),
        'create_photo_form': ReturnPhotoUploadForm(),
    }
    return render(request, 'returns/dashboard.html', context)


@login_required
def appeals_dashboard(request):
    apelaciones = Appeal.objects.select_related('devolucion')
    devoluciones_sin_apelacion = Return.objects.filter(apelaciones__isnull=True)
    todas_devoluciones_sin_apelacion = devoluciones_sin_apelacion.order_by('-fecha_ingreso', '-id')
    q = request.GET.get('q', '')
    status = request.GET.get('status', '')
    apelaciones, period = apply_period_filter(apelaciones, request, 'creado_en__date')

    if q:
        apelaciones = apelaciones.filter(
            Q(numero_apelacion__icontains=q) |
            Q(devolucion__numero_orden__icontains=q) |
            Q(devolucion__sku__icontains=q) |
            Q(devolucion__producto_nombre__icontains=q) |
            Q(estado_cuenta__icontains=q) |
            Q(monto_devuelto__icontains=q)
        )
        devoluciones_sin_apelacion = devoluciones_sin_apelacion.filter(
            Q(numero_orden__icontains=q) |
            Q(sku__icontains=q) |
            Q(ean__icontains=q) |
            Q(producto_nombre__icontains=q) |
            Q(marca__icontains=q) |
            Q(categoria__icontains=q)
        )
    if status:
        apelaciones = apelaciones.filter(status=status)

    def clp_format(value):
        sign = '-' if value < 0 else ''
        return f'{sign}${abs(value):,.0f}'.replace(',', '.') + ' CLP'

    costo_devoluciones_con_pago = 0
    total_pagado_apelaciones = 0
    for apelacion in apelaciones.filter(status='pagado').select_related('devolucion').exclude(monto_devuelto=''):
        if apelacion.monto_pagado is None or apelacion.devolucion.valor_total is None:
            continue
        costo_devoluciones_con_pago += int(apelacion.devolucion.valor_total)
        total_pagado_apelaciones += int(apelacion.monto_pagado)

    diferencia_pagos = costo_devoluciones_con_pago - total_pagado_apelaciones
    perdida_estimada = max(diferencia_pagos, 0)

    stats = {
        'total': apelaciones.count(),
        'en_proceso': apelaciones.filter(status='en_proceso').count(),
        'rechazado': apelaciones.filter(status='rechazado').count(),
        'proceso_pago': apelaciones.filter(status='proceso_pago').count(),
        'pagado': apelaciones.filter(status='pagado').count(),
        'costo_devoluciones_con_pago': costo_devoluciones_con_pago,
        'total_pagado_apelaciones': total_pagado_apelaciones,
        'diferencia_pagos': diferencia_pagos,
        'perdida_estimada': perdida_estimada,
        'costo_devoluciones_con_pago_display': clp_format(costo_devoluciones_con_pago),
        'total_pagado_apelaciones_display': clp_format(total_pagado_apelaciones),
        'diferencia_pagos_display': clp_format(diferencia_pagos),
        'perdida_estimada_display': clp_format(perdida_estimada),
    }

    devoluciones_sin_apelacion_page = paginate_queryset(
        request,
        devoluciones_sin_apelacion.order_by('-fecha_ingreso', '-id'),
        'pendientes_page',
        8,
    )
    apelaciones_page = paginate_queryset(request, apelaciones, 'apelaciones_page', 10)

    context = {
        'apelaciones': apelaciones_page,
        'apelaciones_page': apelaciones_page,
        'apelaciones_page_query': querystring_without(request, 'apelaciones_page'),
        'period': period,
        'period_query': querystring_without(request, 'period', 'apelaciones_page'),
        'devoluciones_sin_apelacion': devoluciones_sin_apelacion_page,
        'devoluciones_sin_apelacion_page': devoluciones_sin_apelacion_page,
        'devoluciones_sin_apelacion_page_query': querystring_without(request, 'pendientes_page'),
        'todas_devoluciones_sin_apelacion': todas_devoluciones_sin_apelacion,
        'stats': stats,
        'q': q,
        'status_filter': status,
        'status_choices': Appeal.STATUS_CHOICES,
        'create_form': AppealForm(),
    }
    return render(request, 'returns/appeals_dashboard.html', context)


@login_required
def report_dashboard(request):
    devoluciones = filtered_report_returns(request)
    if request.GET.get('export') == 'xlsx':
        return report_export_response(devoluciones)
    period, _, _ = period_dates(request)

    def clp_format(value):
        return f'${value:,.0f}'.replace(',', '.') + ' CLP'

    total_valorizado = sum(
        devolucion.precio_venta * devolucion.cantidad
        for devolucion in devoluciones.exclude(precio_venta__isnull=True)
    )
    total_pagado = 0
    for devolucion in devoluciones:
        for apelacion in devolucion.apelaciones.all():
            if apelacion.monto_devuelto and apelacion.monto_devuelto.isdigit():
                if apelacion.status == 'pagado':
                    total_pagado += int(apelacion.monto_devuelto)

    devoluciones_page = paginate_queryset(request, devoluciones, 'page', 15)
    context = {
        'devoluciones': devoluciones_page,
        'devoluciones_page': devoluciones_page,
        'devoluciones_page_query': querystring_without(request, 'page'),
        'period': period,
        'period_query': querystring_without(request, 'period', 'page', 'fecha_desde', 'fecha_hasta'),
        'q': request.GET.get('q', '').strip(),
        'fecha_desde': request.GET.get('fecha_desde', '').strip(),
        'fecha_hasta': request.GET.get('fecha_hasta', '').strip(),
        'seller_filter': request.GET.get('seller', '').strip(),
        'estado_filter': request.GET.get('estado', '').strip(),
        'condicion_filter': request.GET.get('condicion', '').strip(),
        'grado_filter': request.GET.get('grado', '').strip(),
        'apelacion_filter': request.GET.get('apelacion', '').strip(),
        'appeal_status_filter': request.GET.get('appeal_status', '').strip(),
        'seller_choices': Return.SELLER_CHOICES,
        'estado_choices': Return.ESTADO_CHOICES,
        'condicion_choices': Return.CONDICION_CHOICES,
        'grado_choices': Return.GRADO_CHOICES,
        'appeal_status_choices': Appeal.STATUS_CHOICES,
        'total_filtrado': devoluciones.count(),
        'total_valorizado_display': clp_format(total_valorizado),
        'total_pagado_display': clp_format(total_pagado),
        'con_apelacion': devoluciones.filter(apelaciones__isnull=False).count(),
    }
    return render(request, 'returns/report.html', context)


@staff_member_required(login_url='login')
def users_dashboard(request):
    if request.method == 'POST':
        form = PlatformUserCreationForm(request.POST)
        if form.is_valid():
            usuario = form.save()
            messages.success(request, f'Usuario {usuario.username} creado correctamente.')
            return redirect('users_dashboard')
    else:
        form = PlatformUserCreationForm()

    q = request.GET.get('q', '').strip()
    for usuario_existente in User.objects.all().only('id'):
        UserProfile.objects.get_or_create(user=usuario_existente)

    usuarios = User.objects.order_by('-date_joined', 'username')
    if q:
        usuarios = usuarios.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q)
        )
    usuarios_page = paginate_queryset(request, usuarios, 'page', 12)
    context = {
        'form': form,
        'usuarios': usuarios_page,
        'usuarios_page': usuarios_page,
        'usuarios_page_query': querystring_without(request, 'page'),
        'q': q,
    }
    return render(request, 'returns/users_dashboard.html', context)


@staff_member_required(login_url='login')
def user_edit(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = PlatformUserUpdateForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, f'Usuario {usuario.username} actualizado correctamente.')
            return redirect('users_dashboard')
    else:
        form = PlatformUserUpdateForm(instance=usuario)
    return render(request, 'returns/user_form.html', {'form': form, 'usuario': usuario})


@staff_member_required(login_url='login')
def user_delete(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if usuario.pk == request.user.pk:
        messages.error(request, 'No puedes eliminar tu propio usuario mientras estás conectado.')
        return redirect('users_dashboard')

    if request.method == 'POST':
        username = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario {username} eliminado correctamente.')
        return redirect('users_dashboard')
    return render(request, 'returns/user_confirm_delete.html', {'usuario': usuario})


@login_required
def appeal_create(request):
    if request.method == 'POST':
        form = AppealForm(request.POST)
        if form.is_valid():
            apelacion = form.save(commit=False)
            apelacion.creado_por = request.user
            apelacion.actualizado_por = request.user
            apelacion.save()
            messages.success(request, 'Apelación creada correctamente.')
            return redirect('appeal_detail', pk=apelacion.pk)
    else:
        form = AppealForm(initial={'devolucion': request.GET.get('devolucion')})
    return render(request, 'returns/appeal_form.html', {'form': form, 'action': 'Crear'})


@login_required
def appeal_detail(request, pk):
    apelacion = get_object_or_404(Appeal.objects.select_related('devolucion'), pk=pk)
    return render(request, 'returns/appeal_detail.html', {'apelacion': apelacion})


@login_required
def appeal_update_status(request, pk):
    apelacion = get_object_or_404(Appeal, pk=pk)
    if request.method == 'POST':
        status = request.POST.get('status')
        valid_statuses = dict(Appeal.STATUS_CHOICES)
        if status in valid_statuses:
            monto_devuelto = request.POST.get('monto_devuelto', '').strip().replace('$', '').replace('.', '').replace(' ', '')
            if status == 'pagado' and not monto_devuelto:
                messages.error(request, 'Ingresa el monto efectivamente devuelto en $ CLP para marcar la apelación como pagada.')
                return redirect('appeal_detail', pk=apelacion.pk)
            if monto_devuelto and not monto_devuelto.isdigit():
                messages.error(request, 'Ingresa un monto válido en pesos chilenos.')
                return redirect('appeal_detail', pk=apelacion.pk)

            apelacion.status = status
            apelacion.actualizado_por = request.user
            update_fields = ['status', 'actualizado_por', 'actualizado_en']
            if status == 'pagado':
                apelacion.monto_devuelto = monto_devuelto
                update_fields.append('monto_devuelto')
            apelacion.save(update_fields=update_fields)
            messages.success(request, f'Apelación actualizada a {valid_statuses[status]}.')
        else:
            messages.error(request, 'Status inválido.')
    return redirect('appeal_detail', pk=apelacion.pk)


@login_required
def appeal_edit(request, pk):
    apelacion = get_object_or_404(Appeal, pk=pk)
    if request.method == 'POST':
        form = AppealForm(request.POST, instance=apelacion)
        if form.is_valid():
            apelacion = form.save(commit=False)
            apelacion.actualizado_por = request.user
            apelacion.save()
            messages.success(request, 'Apelación actualizada correctamente.')
            return redirect('appeal_detail', pk=apelacion.pk)
    else:
        form = AppealForm(instance=apelacion)
    return render(request, 'returns/appeal_form.html', {'form': form, 'action': 'Editar', 'apelacion': apelacion})


@login_required
def appeal_delete(request, pk):
    apelacion = get_object_or_404(Appeal, pk=pk)
    if request.method == 'POST':
        apelacion.delete()
        messages.success(request, 'Apelación eliminada.')
        return redirect('appeals_dashboard')
    return render(request, 'returns/appeal_confirm_delete.html', {'apelacion': apelacion})


@login_required
def return_create(request):
    if request.method == 'POST':
        form = ReturnForm(request.POST)
        photo_form = ReturnPhotoUploadForm(request.POST, request.FILES)
        if form.is_valid() and photo_form.is_valid():
            devolucion = form.save(commit=False)
            devolucion.creado_por = request.user
            devolucion.save()
            for foto in photo_form.cleaned_data['fotos']:
                ReturnPhoto.objects.create(devolucion=devolucion, foto=foto)
            messages.success(request, f'Devolución #{devolucion.id} creada correctamente.')
            return redirect('return_detail', pk=devolucion.pk)
    else:
        form = ReturnForm()
        photo_form = ReturnPhotoUploadForm()
    return render(request, 'returns/return_form.html', {'form': form, 'photo_form': photo_form, 'action': 'Crear'})


@login_required
def return_detail(request, pk):
    devolucion = get_object_or_404(Return.objects.prefetch_related('apelaciones'), pk=pk)
    return render(request, 'returns/return_detail.html', {'devolucion': devolucion})


@login_required
def return_edit(request, pk):
    devolucion = get_object_or_404(Return, pk=pk)
    if request.method == 'POST':
        form = ReturnForm(request.POST, instance=devolucion)
        delete_photo_ids = request.POST.getlist('delete_photos')
        photo_form = ReturnPhotoUploadForm(request.POST, request.FILES)
        if form.is_valid() and photo_form.is_valid():
            form.save()
            devolucion.fotos.filter(pk__in=delete_photo_ids).delete()
            for foto in photo_form.cleaned_data['fotos']:
                ReturnPhoto.objects.create(devolucion=devolucion, foto=foto)
            messages.success(request, 'Devolución actualizada correctamente.')
            return redirect('return_detail', pk=devolucion.pk)
    else:
        form = ReturnForm(instance=devolucion)
        photo_form = ReturnPhotoUploadForm()
    return render(request, 'returns/return_form.html', {
        'form': form,
        'photo_form': photo_form,
        'action': 'Editar',
        'devolucion': devolucion,
    })


@login_required
def return_delete(request, pk):
    devolucion = get_object_or_404(Return, pk=pk)
    if request.method == 'POST':
        devolucion.delete()
        messages.success(request, 'Devolución eliminada.')
        return redirect('returns_dashboard')
    return render(request, 'returns/return_confirm_delete.html', {'devolucion': devolucion})
