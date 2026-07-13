from django.contrib import admin
from django import forms
from django.conf import settings
from django.core.management import call_command
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, time
from io import StringIO
from pathlib import Path
import json
import tempfile
import unicodedata
import zipfile
from .models import Appeal, Product, Return, ReturnPhoto, SiteConfiguration


class FullBackupImportForm(forms.Form):
    archivo = forms.FileField(
        label='Archivo backup ZIP',
        help_text='Sube un backup .zip generado desde esta misma herramienta.',
    )

    def clean_archivo(self):
        archivo = self.cleaned_data['archivo']
        if not archivo.name.lower().endswith('.zip'):
            raise forms.ValidationError('Sube un archivo .zip de backup.')
        return archivo


@admin.register(SiteConfiguration)
class SiteConfigurationAdmin(admin.ModelAdmin):
    fields = ['timezone', 'actualizado_en']
    readonly_fields = ['actualizado_en']
    change_list_template = 'admin/returns/siteconfiguration/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'backup/',
                self.admin_site.admin_view(self.backup_view),
                name='returns_full_backup',
            ),
            path(
                'backup/exportar/',
                self.admin_site.admin_view(self.export_backup_view),
                name='returns_full_backup_export',
            ),
            path(
                'limpiar-datos/',
                self.admin_site.admin_view(self.clear_work_data_view),
                name='returns_clear_work_data',
            ),
        ]
        return custom_urls + urls

    def has_add_permission(self, request):
        return not SiteConfiguration.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    def backup_view(self, request):
        if request.method == 'POST':
            form = FullBackupImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    self.import_backup(form.cleaned_data['archivo'])
                except (KeyError, ValueError, zipfile.BadZipFile) as error:
                    self.message_user(request, f'No se pudo importar el backup: {error}', level='ERROR')
                    return redirect('admin:returns_full_backup')

                self.message_user(request, 'Backup importado correctamente.')
                return redirect('admin:returns_full_backup')
        else:
            form = FullBackupImportForm()

        return render(
            request,
            'admin/returns/siteconfiguration/backup.html',
            {
                'form': form,
                'title': 'Backup completo',
                'opts': self.model._meta,
            },
        )

    def export_backup_view(self, request):
        timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="backup_devoluciones_{timestamp}.zip"'

        data_output = StringIO()
        call_command(
            'dumpdata',
            'auth.User',
            'auth.Group',
            'returns',
            format='json',
            indent=2,
            stdout=data_output,
        )
        manifest = {
            'created_at': timezone.localtime().isoformat(),
            'format': 'devoluciones-full-backup-v1',
            'contains': ['data.json', 'media/'],
        }

        media_root = Path(settings.MEDIA_ROOT)
        with zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED) as backup_zip:
            backup_zip.writestr('manifest.json', json.dumps(manifest, ensure_ascii=False, indent=2))
            backup_zip.writestr('data.json', data_output.getvalue())
            if media_root.exists():
                for path in media_root.rglob('*'):
                    if path.is_file():
                        backup_zip.write(path, Path('media') / path.relative_to(media_root))

        return response

    def import_backup(self, archivo):
        media_root = Path(settings.MEDIA_ROOT)
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            with zipfile.ZipFile(archivo) as backup_zip:
                names = backup_zip.namelist()
                if 'data.json' not in names:
                    raise ValueError('el ZIP no contiene data.json')

                for member in names:
                    target = temp_path / member
                    if not str(target.resolve()).startswith(str(temp_path.resolve())):
                        raise ValueError('el ZIP contiene rutas no permitidas')
                backup_zip.extractall(temp_path)

            call_command('loaddata', str(temp_path / 'data.json'))

            extracted_media = temp_path / 'media'
            if extracted_media.exists():
                media_root.mkdir(parents=True, exist_ok=True)
                for source in extracted_media.rglob('*'):
                    if source.is_file():
                        target = media_root / source.relative_to(extracted_media)
                        target.parent.mkdir(parents=True, exist_ok=True)
                        target.write_bytes(source.read_bytes())

    def clear_work_data_view(self, request):
        if request.method == 'POST':
            counts = {
                'devoluciones': Return.objects.count(),
                'apelaciones': Appeal.objects.count(),
                'fotos': ReturnPhoto.objects.count(),
            }
            Return.objects.all().delete()
            self.message_user(
                request,
                (
                    'Datos de trabajo eliminados: '
                    f"{counts['devoluciones']} devoluciones, "
                    f"{counts['apelaciones']} apelaciones, "
                    f"{counts['fotos']} fotos."
                ),
            )
            return redirect('admin:returns_siteconfiguration_changelist')

        return render(
            request,
            'admin/returns/siteconfiguration/clear_work_data.html',
            {
                'title': 'Eliminar datos de trabajo',
                'opts': self.model._meta,
                'returns_count': Return.objects.count(),
                'appeals_count': Appeal.objects.count(),
                'photos_count': ReturnPhoto.objects.count(),
                'products_count': Product.objects.count(),
            },
        )


class ProductImportForm(forms.Form):
    archivo = forms.FileField(
        label='Archivo Excel',
        help_text='Usa una planilla .xlsx con columnas: sku, nombre, ean, marca, categoria.',
    )

    def clean_archivo(self):
        archivo = self.cleaned_data['archivo']
        if not archivo.name.lower().endswith('.xlsx'):
            raise forms.ValidationError('Sube un archivo Excel con extension .xlsx.')
        return archivo


class ReturnImportForm(forms.Form):
    archivo = forms.FileField(
        label='Archivo Excel',
        help_text='Usa una planilla .xlsx con columnas: numero_orden, fecha_ingreso, seller, sku, ean, producto_nombre, marca, categoria, precio_venta, cantidad, ingresado_bodega, estado, condicion_producto, sub_danos, detalles_dano y notas_internas.',
    )

    def clean_archivo(self):
        archivo = self.cleaned_data['archivo']
        if not archivo.name.lower().endswith('.xlsx'):
            raise forms.ValidationError('Sube un archivo Excel con extension .xlsx.')
        return archivo


class AppealImportForm(forms.Form):
    archivo = forms.FileField(
        label='Archivo Excel',
        help_text='Usa una planilla .xlsx con columnas: numero_orden, numero_apelacion, fecha_apelacion, detalle, status, monto_apelado y monto_devuelto.',
    )

    def clean_archivo(self):
        archivo = self.cleaned_data['archivo']
        if not archivo.name.lower().endswith('.xlsx'):
            raise forms.ValidationError('Sube un archivo Excel con extension .xlsx.')
        return archivo


def clean_excel_value(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_excel_key(value):
    value = unicodedata.normalize('NFKD', clean_excel_value(value))
    value = ''.join(char for char in value if not unicodedata.combining(char))
    return value.lower().replace(' ', '_').replace('-', '_')


def build_header_map(headers):
    return {
        normalize_excel_key(header): index
        for index, header in enumerate(headers)
        if clean_excel_value(header)
    }


def excel_row_value(row, header_map, *names):
    for name in names:
        index = header_map.get(normalize_excel_key(name))
        if index is not None and index < len(row):
            return clean_excel_value(row[index])
    return ''


def excel_raw_value(row, header_map, *names):
    for name in names:
        index = header_map.get(normalize_excel_key(name))
        if index is not None and index < len(row):
            return row[index]
    return None


def excel_date_value(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw_value = clean_excel_value(value)
    for date_format in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(raw_value, date_format).date()
        except ValueError:
            continue
    return None


def excel_datetime_value(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        parsed_value = value
    else:
        parsed_date = excel_date_value(value)
        if parsed_date is None:
            return None
        parsed_value = datetime.combine(parsed_date, time.min)

    if timezone.is_naive(parsed_value):
        return timezone.make_aware(parsed_value, timezone.get_current_timezone())
    return parsed_value


def excel_choice_value(value, choices):
    value = normalize_excel_key(value)
    if not value:
        return ''
    for choice_value, choice_label in choices:
        if value in {normalize_excel_key(choice_value), normalize_excel_key(choice_label)}:
            return choice_value
    return ''


def excel_bool_value(value):
    value = normalize_excel_key(value)
    return value in {'1', 'si', 'sí', 'true', 'verdadero', 'yes', 'y', 'ingresado', 'ingresada'}


def excel_int_value(value, default=1):
    value = clean_excel_value(value)
    if not value:
        return default
    try:
        return max(int(float(value)), 1)
    except (TypeError, ValueError):
        return default


def excel_decimal_value(value):
    value = clean_excel_value(value)
    if not value:
        return None
    value = value.replace('$', '').replace('CLP', '').replace('clp', '').replace(' ', '')
    if ',' in value:
        value = value.replace('.', '').replace(',', '.')
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return None


def excel_sub_danos_value(value):
    labels = Return.SUB_DANO_CHOICES
    values = []
    raw_items = clean_excel_value(value).replace('|', ',').replace(';', ',').split(',')
    for item in raw_items:
        sub_dano = excel_choice_value(item, labels)
        if sub_dano and sub_dano not in values:
            values.append(sub_dano)
    return values


def style_excel_header(sheet):
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = 'A2'


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ['sku', 'nombre', 'ean', 'marca', 'categoria', 'actualizado_en']
    search_fields = ['sku', 'nombre', 'ean', 'marca', 'categoria']
    list_filter = ['marca', 'categoria']
    readonly_fields = ['creado_en', 'actualizado_en']
    change_list_template = 'admin/returns/product/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'importar-excel/',
                self.admin_site.admin_view(self.import_excel_view),
                name='returns_product_import_excel',
            ),
            path(
                'plantilla-excel/',
                self.admin_site.admin_view(self.download_excel_template_view),
                name='returns_product_excel_template',
            ),
        ]
        return custom_urls + urls

    def download_excel_template_view(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self.message_user(
                request,
                'Para descargar la plantilla instala openpyxl: pip install openpyxl',
                level='ERROR',
            )
            return redirect('admin:returns_product_changelist')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Productos'
        sheet.append(['sku', 'nombre', 'ean', 'marca', 'categoria'])
        sheet.append(['SKU-12345', 'Nombre del producto', '7800000000000', 'Marca ejemplo', 'Categoría ejemplo'])

        header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 38
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 22
        sheet.column_dimensions['E'].width = 24
        sheet.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
        workbook.save(response)
        return response

    def import_excel_view(self, request):
        if request.method == 'POST':
            form = ProductImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    self.message_user(
                        request,
                        'Para importar Excel instala openpyxl: pip install openpyxl',
                        level='ERROR',
                    )
                    return redirect('admin:returns_product_changelist')

                workbook = load_workbook(form.cleaned_data['archivo'], read_only=True, data_only=True)
                sheet = workbook.active
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, None)

                if not headers:
                    self.message_user(request, 'El archivo no tiene encabezados.', level='ERROR')
                    return redirect('admin:returns_product_import_excel')

                header_map = {
                    clean_excel_value(header).lower().replace(' ', '_'): index
                    for index, header in enumerate(headers)
                }
                sku_index = header_map.get('sku')
                nombre_index = (
                    header_map.get('nombre')
                    if 'nombre' in header_map
                    else header_map.get(
                        'nombre_del_producto',
                        header_map.get('nombre_producto', header_map.get('producto')),
                    )
                )
                ean_index = header_map.get('ean')
                marca_index = header_map.get('marca')
                categoria_index = header_map.get('categoria', header_map.get('categoría'))

                if sku_index is None or nombre_index is None:
                    self.message_user(
                        request,
                        'El Excel debe incluir al menos las columnas "sku" y "nombre".',
                        level='ERROR',
                    )
                    return redirect('admin:returns_product_import_excel')

                created_count = 0
                updated_count = 0
                skipped_count = 0

                for row in rows:
                    sku = clean_excel_value(row[sku_index] if sku_index < len(row) else '')
                    nombre = clean_excel_value(row[nombre_index] if nombre_index < len(row) else '')
                    ean = clean_excel_value(row[ean_index] if ean_index is not None and ean_index < len(row) else '')
                    marca = clean_excel_value(row[marca_index] if marca_index is not None and marca_index < len(row) else '')
                    categoria = clean_excel_value(row[categoria_index] if categoria_index is not None and categoria_index < len(row) else '')

                    if not sku or not nombre:
                        skipped_count += 1
                        continue

                    _, created = Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            'nombre': nombre,
                            'ean': ean,
                            'marca': marca,
                            'categoria': categoria,
                        },
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                self.message_user(
                    request,
                    f'Importación lista: {created_count} creados, {updated_count} actualizados, {skipped_count} omitidos.',
                )
                workbook.close()
                return redirect('admin:returns_product_changelist')
        else:
            form = ProductImportForm()

        return render(
            request,
            'admin/returns/product/import_excel.html',
            {
                'form': form,
                'title': 'Importar productos desde Excel',
                'opts': self.model._meta,
            },
        )


class ReturnPhotoInline(admin.TabularInline):
    model = ReturnPhoto
    extra = 0


@admin.register(Return)
class ReturnAdmin(admin.ModelAdmin):
    list_display = ['id', 'numero_orden', 'seller_display', 'sku', 'ean', 'producto_nombre', 'marca', 'categoria', 'cantidad', 'precio_venta', 'ingresado_bodega', 'estado', 'condicion_producto', 'grado', 'fecha_ingreso', 'creado_por']
    list_filter = ['seller', 'ingresado_bodega', 'estado', 'condicion_producto', 'grado', 'marca', 'categoria', 'fecha_ingreso']
    search_fields = ['numero_orden', 'seller', 'sku', 'ean', 'producto_nombre', 'marca', 'categoria']
    inlines = [ReturnPhotoInline]
    readonly_fields = ['estado', 'grado', 'fecha_ingreso', 'fecha_actualizacion', 'creado_por']
    change_list_template = 'admin/returns/return/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'importar-excel/',
                self.admin_site.admin_view(self.import_excel_view),
                name='returns_return_import_excel',
            ),
            path(
                'plantilla-excel/',
                self.admin_site.admin_view(self.download_excel_template_view),
                name='returns_return_excel_template',
            ),
        ]
        return custom_urls + urls

    def download_excel_template_view(self, request):
        try:
            from openpyxl import Workbook
        except ImportError:
            self.message_user(request, 'Para descargar la plantilla instala openpyxl: pip install openpyxl', level='ERROR')
            return redirect('admin:returns_return_changelist')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Devoluciones'
        headers = [
            'numero_orden',
            'fecha_ingreso',
            'seller',
            'sku',
            'ean',
            'producto_nombre',
            'marca',
            'categoria',
            'precio_venta',
            'cantidad',
            'ingresado_bodega',
            'estado',
            'condicion_producto',
            'sub_danos',
            'detalles_dano',
            'notas_internas',
        ]
        sheet.append(headers)
        sheet.append([
            'ORD-2026-001',
            '13/07/2026',
            'falabella',
            'SKU-12345',
            '7800000000000',
            'Nombre del producto',
            'Marca ejemplo',
            'Categoria ejemplo',
            '49990',
            '1',
            'si',
            'recibido',
            'nuevo',
            'rayado, sin_accesorios',
            'Detalle del daño',
            'Nota interna',
        ])
        style_excel_header(sheet)
        widths = [18, 16, 16, 18, 18, 34, 20, 22, 14, 10, 18, 16, 22, 28, 34, 34]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_devoluciones.xlsx"'
        workbook.save(response)
        return response

    def import_excel_view(self, request):
        if request.method == 'POST':
            form = ReturnImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    self.message_user(request, 'Para importar Excel instala openpyxl: pip install openpyxl', level='ERROR')
                    return redirect('admin:returns_return_changelist')

                workbook = load_workbook(form.cleaned_data['archivo'], read_only=True, data_only=True)
                sheet = workbook.active
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, None)
                if not headers:
                    self.message_user(request, 'El archivo no tiene encabezados.', level='ERROR')
                    return redirect('admin:returns_return_import_excel')

                header_map = build_header_map(headers)
                required_headers = {'numero_orden', 'seller', 'sku'}
                if not required_headers.issubset(header_map):
                    self.message_user(request, 'El Excel debe incluir las columnas "numero_orden", "seller" y "sku".', level='ERROR')
                    return redirect('admin:returns_return_import_excel')

                created_count = 0
                updated_count = 0
                skipped_count = 0
                with transaction.atomic():
                    for row in rows:
                        numero_orden = excel_row_value(row, header_map, 'numero_orden', 'número_orden', 'orden')
                        seller = excel_choice_value(excel_row_value(row, header_map, 'seller'), Return.SELLER_CHOICES)
                        sku = excel_row_value(row, header_map, 'sku')
                        fecha_ingreso_raw = excel_raw_value(row, header_map, 'fecha_ingreso', 'fecha')
                        fecha_ingreso = excel_date_value(fecha_ingreso_raw)
                        if not numero_orden or not seller or not sku or (clean_excel_value(fecha_ingreso_raw) and fecha_ingreso is None):
                            skipped_count += 1
                            continue

                        product = Product.objects.filter(sku__iexact=sku).first()
                        ean = excel_row_value(row, header_map, 'ean') or (product.ean if product else '')
                        producto_nombre = excel_row_value(row, header_map, 'producto_nombre', 'nombre', 'producto') or (product.nombre if product else '')
                        marca = excel_row_value(row, header_map, 'marca') or (product.marca if product else '')
                        categoria = excel_row_value(row, header_map, 'categoria', 'categoría') or (product.categoria if product else '')
                        condicion_producto = excel_choice_value(
                            excel_row_value(row, header_map, 'condicion_producto', 'estado_producto'),
                            Return.CONDICION_CHOICES,
                        ) or 'nuevo'
                        estado = excel_choice_value(excel_row_value(row, header_map, 'estado'), Return.ESTADO_CHOICES) or 'recibido'

                        devolucion, created = Return.objects.get_or_create(
                            numero_orden=numero_orden,
                            defaults={
                                'seller': seller,
                                'sku': sku,
                                'creado_por': request.user,
                            },
                        )
                        devolucion.seller = seller
                        devolucion.sku = sku
                        devolucion.ean = ean
                        devolucion.producto_nombre = producto_nombre
                        devolucion.marca = marca
                        devolucion.categoria = categoria
                        devolucion.precio_venta = excel_decimal_value(excel_row_value(row, header_map, 'precio_venta', 'precio'))
                        devolucion.cantidad = excel_int_value(excel_row_value(row, header_map, 'cantidad'), default=1)
                        devolucion.ingresado_bodega = excel_bool_value(excel_row_value(row, header_map, 'ingresado_bodega', 'bodega'))
                        devolucion.estado = estado
                        devolucion.condicion_producto = condicion_producto
                        devolucion.sub_danos = excel_sub_danos_value(excel_row_value(row, header_map, 'sub_danos', 'sub_daños'))
                        devolucion.detalles_dano = excel_row_value(row, header_map, 'detalles_dano', 'detalles_daño')
                        devolucion.notas_internas = excel_row_value(row, header_map, 'notas_internas')
                        devolucion.save()
                        if fecha_ingreso is not None:
                            Return.objects.filter(pk=devolucion.pk).update(fecha_ingreso=fecha_ingreso)

                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                workbook.close()
                self.message_user(request, f'Importación lista: {created_count} creadas, {updated_count} actualizadas, {skipped_count} omitidas.')
                return redirect('admin:returns_return_changelist')
        else:
            form = ReturnImportForm()

        return render(
            request,
            'admin/returns/return/import_excel.html',
            {
                'form': form,
                'title': 'Importar devoluciones desde Excel',
                'opts': self.model._meta,
            },
        )

    @admin.display(description='Seller', ordering='seller')
    def seller_display(self, obj):
        return obj.get_seller_display()


@admin.register(ReturnPhoto)
class ReturnPhotoAdmin(admin.ModelAdmin):
    list_display = ['id', 'devolucion', 'descripcion', 'subida_en']


@admin.register(Appeal)
class AppealAdmin(admin.ModelAdmin):
    list_display = ['id', 'numero_apelacion', 'orden_display', 'status', 'estado_cuenta', 'monto_devuelto', 'creado_en', 'creado_por', 'actualizado_en', 'actualizado_por']
    list_filter = ['status', 'creado_en']
    search_fields = ['numero_apelacion', 'devolucion__numero_orden', 'devolucion__sku', 'devolucion__producto_nombre', 'estado_cuenta', 'monto_devuelto']
    readonly_fields = ['creado_en', 'actualizado_en', 'creado_por', 'actualizado_por']
    change_list_template = 'admin/returns/appeal/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'importar-excel/',
                self.admin_site.admin_view(self.import_excel_view),
                name='returns_appeal_import_excel',
            ),
            path(
                'plantilla-excel/',
                self.admin_site.admin_view(self.download_excel_template_view),
                name='returns_appeal_excel_template',
            ),
        ]
        return custom_urls + urls

    def download_excel_template_view(self, request):
        try:
            from openpyxl import Workbook
        except ImportError:
            self.message_user(request, 'Para descargar la plantilla instala openpyxl: pip install openpyxl', level='ERROR')
            return redirect('admin:returns_appeal_changelist')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Apelaciones'
        sheet.append(['numero_orden', 'numero_apelacion', 'fecha_apelacion', 'detalle', 'status', 'monto_apelado', 'monto_devuelto'])
        sheet.append(['ORD-2026-001', 'TICKET-001', '13/07/2026', 'Detalle de la apelacion', 'en_proceso', '25990', ''])
        style_excel_header(sheet)
        widths = [18, 22, 18, 42, 18, 16, 16]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_apelaciones.xlsx"'
        workbook.save(response)
        return response

    def import_excel_view(self, request):
        if request.method == 'POST':
            form = AppealImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    self.message_user(request, 'Para importar Excel instala openpyxl: pip install openpyxl', level='ERROR')
                    return redirect('admin:returns_appeal_changelist')

                workbook = load_workbook(form.cleaned_data['archivo'], read_only=True, data_only=True)
                sheet = workbook.active
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, None)
                if not headers:
                    self.message_user(request, 'El archivo no tiene encabezados.', level='ERROR')
                    return redirect('admin:returns_appeal_import_excel')

                header_map = build_header_map(headers)
                required_headers = {'numero_orden', 'numero_apelacion', 'detalle'}
                if not required_headers.issubset(header_map):
                    self.message_user(request, 'El Excel debe incluir las columnas "numero_orden", "numero_apelacion" y "detalle".', level='ERROR')
                    return redirect('admin:returns_appeal_import_excel')

                created_count = 0
                updated_count = 0
                skipped_count = 0
                with transaction.atomic():
                    for row in rows:
                        numero_orden = excel_row_value(row, header_map, 'numero_orden', 'número_orden', 'orden')
                        numero_apelacion = excel_row_value(row, header_map, 'numero_apelacion', 'número_apelación', 'ticket')
                        detalle = excel_row_value(row, header_map, 'detalle')
                        fecha_apelacion_raw = excel_raw_value(row, header_map, 'fecha_apelacion', 'creado_en', 'fecha_creacion', 'fecha')
                        fecha_apelacion = excel_datetime_value(fecha_apelacion_raw)
                        devolucion = Return.objects.filter(numero_orden=numero_orden).first()
                        if not devolucion or not numero_apelacion or not detalle or (clean_excel_value(fecha_apelacion_raw) and fecha_apelacion is None):
                            skipped_count += 1
                            continue

                        status = excel_choice_value(excel_row_value(row, header_map, 'status', 'estado'), Appeal.STATUS_CHOICES) or 'en_proceso'
                        estado_cuenta = clean_excel_value(excel_row_value(row, header_map, 'monto_apelado', 'estado_cuenta', 'monto')).replace('$', '').replace('.', '').replace(' ', '')
                        monto_devuelto = clean_excel_value(excel_row_value(row, header_map, 'monto_devuelto', 'monto_pagado')).replace('$', '').replace('.', '').replace(' ', '')
                        apelacion, created = Appeal.objects.get_or_create(
                            devolucion=devolucion,
                            numero_apelacion=numero_apelacion,
                            defaults={
                                'detalle': detalle,
                                'creado_por': request.user,
                            },
                        )
                        apelacion.detalle = detalle
                        apelacion.status = status
                        apelacion.estado_cuenta = estado_cuenta
                        apelacion.monto_devuelto = monto_devuelto
                        if created:
                            apelacion.creado_por = request.user
                        else:
                            apelacion.actualizado_por = request.user
                        apelacion.save()
                        if fecha_apelacion is not None:
                            Appeal.objects.filter(pk=apelacion.pk).update(creado_en=fecha_apelacion)

                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                workbook.close()
                self.message_user(request, f'Importación lista: {created_count} creadas, {updated_count} actualizadas, {skipped_count} omitidas.')
                return redirect('admin:returns_appeal_changelist')
        else:
            form = AppealImportForm()

        return render(
            request,
            'admin/returns/appeal/import_excel.html',
            {
                'form': form,
                'title': 'Importar apelaciones desde Excel',
                'opts': self.model._meta,
            },
        )

    @admin.display(description='Orden', ordering='devolucion__numero_orden')
    def orden_display(self, obj):
        return obj.devolucion.numero_orden
